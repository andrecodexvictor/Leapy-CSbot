import openpyxl
import os
import re

def compile_database():
    excel_path = r"docs/leapy-dados-ficticios-rag.xlsx"
    db_ts_path = r"server/db.ts"
    
    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} not found.")
        return

    wb = openpyxl.load_workbook(excel_path)
    sheet = wb["Chunks"]
    
    # Read headers
    headers = [cell.value for cell in sheet[1]]
    
    # Read rows
    chunks = []
    for r in range(2, sheet.max_row + 1):
        row_vals = [sheet.cell(r, c).value for c in range(1, len(headers) + 1)]
        if not any(val is not None for val in row_vals):
            continue
        chunk_dict = dict(zip(headers, row_vals))
        chunks.append(chunk_dict)
        
    print(f"Read {len(chunks)} chunks from excel.")
    
    nodes_code = []
    edges_code = []
    
    # 1. Add static Concept Nodes
    concepts = [
        {
            "id": 'cota_aprendiz',
            "title": 'Cota de Jovem Aprendiz',
            "type": 'concept',
            "description": 'Regras de obrigatoriedade legal de 5% a 15% de contratação de jovens aprendizes conforme CLT.',
            "keywords": ['aprendiz', 'jovem aprendiz', 'cota', 'clt', 'lei', 'obrigatoriedade', 'calculo']
        },
        {
            "id": 'cota_pcd',
            "title": 'Cota de PCD',
            "type": 'concept',
            "description": 'Cota obrigatória de contratação de pessoas com deviciência (2% a 5%) aplicável a empresas a partir de 100 funcionários.',
            "keywords": ['pcd', 'deficiente', 'cota', 'lei', 'obrigatoriedade', '100']
        },
        {
            "id": 'elegibilidade_clt',
            "title": 'Elegibilidade CLT',
            "type": 'concept',
            "description": 'Políticas de elegibilidade para planos de saúde, vale refeição premium e coparticipação para contratos CLT.',
            "keywords": ['clt', 'elegibilidade', 'beneficio', 'plano de saude', 'sulamerica', 'experiencia']
        },
        {
            "id": 'elegibilidade_estagio',
            "title": 'Elegibilidade Estágio',
            "type": 'concept',
            "description": 'Direitos limitados de estagiários na Leapy: exclusivo vale refeição reduzido e seguro de vida, sem direito a plano de saúde ou Gympass.',
            "keywords": ['estagiario', 'estagio', 'elegibilidade', 'bolsa', 'beneficio', 'vr']
        },
        {
            "id": 'operacao_sudeste',
            "title": 'Homologação Regional (Sudeste e PR)',
            "type": 'concept',
            "description": 'Automação completa homologada de convenções coletivas e dissídios retroativos restrita a SP, RJ, MG e PR.',
            "keywords": ['sp', 'rj', 'mg', 'pr', 'homologação', 'dissidio', 'sindicato', 'cct']
        },
        {
            "id": 'operacao_nacional',
            "title": 'Abrangência Nacional',
            "type": 'concept',
            "description": 'Capacidade técnica de rodar folha em todo o Brasil com reajustes manuais nas demais regiões.',
            "keywords": ['nacional', 'brasil', 'territorio', 'manual', 'customizado']
        },
        {
            "id": 'portal_colaborador',
            "title": 'Portal do Colaborador Self-Service',
            "type": 'concept',
            "description": 'Interface de autoatendimento para holerites assinados, envio de atestados e férias sem gargalos de RH.',
            "keywords": ['portal', 'self service', 'holerite', 'atestado', 'colaborador', 'funcionario']
        },
        {
            "id": 'solicitacao_ferias',
            "title": 'Fluxo de Férias',
            "type": 'concept',
            "description": 'Regra estrita de 30 dias de antecedência mínima e aprovação sequencial de Gestor e RH, com cancelamento automático se atrasar.',
            "keywords": ['ferias', 'antecedencia', 'solicitação', 'gestor', 'rh', 'cancelamento', 'aprovação']
        },
        {
            "id": 'transicao_estagio',
            "title": 'Transição e Efetivação CLT',
            "type": 'concept',
            "description": 'Regras para efetivar estagiários: antecedência de 15 dias, notas médias de desempenho > 7.5, e irredutibilidade do salário.',
            "keywords": ['efetivação', 'efetivar', 'transição', 'estagiario', 'desempenho', 'nota', 'salario']
        },
        {
            "id": 'integracao_erp',
            "title": 'Integração de Sistemas',
            "type": 'concept',
            "description": 'Solução para sistemas legados como Totvs, Sênior e SAP via API RESTful ou exportadores personalizados agendados.',
            "keywords": ['integracao', 'erp', 'senior', 'totvs', 'sap', 'api', 'csv']
        },
        {
            "id": 'seguranca_lgpd',
            "title": 'Segurança e LGPD',
            "type": 'concept',
            "description": 'Criptografia robusta TLS 1.3 em trânsito e AES-256 em repouso com controle baseado em perfil (RBAC) conforme LGPD.',
            "keywords": ['segurança', 'lgpd', 'criptografia', 'dados', 'privacidade', 'aes', 'tls']
        }
    ]
    
    # 2. Add Chunk Document Nodes
    ENRICHED_TEXTS = {
        "CHK-001": "A Leapy é uma plataforma pioneira que apoia empresas na gestão completa do programa de jovens aprendizes de ponta a ponta. Ela combina a formação teórica obrigatória (através da escola técnica certificada Leapy GO) com uma infraestrutura de software SaaS para automação operacional (assinaturas de contratos, relatórios de desenvolvimento, controle de ponto e integração com ERPs de folha de pagamento). O objetivo da Leapy é eliminar a carga burocrática das empresas, permitindo que a cota legal de aprendizagem seja utilizada de forma estratégica como um celeiro de novos talentos de alta performance.",
        "CHK-002": "A operação da Leapy está estruturada em quatro frentes principais: 1) Estruturação e Acompanhamento: triagem socioeconômica, recrutamento dinâmico e integração rápida do jovem à cultura do cliente. 2) Formação Teórica Obrigatória: aulas ministradas pela escola técnica oficial Leapy GO nas áreas de tecnologia (análise de dados, Excel avançado, programação e IA) e soft skills. 3) Plataforma SaaS centralizada: painel para acompanhamento de notas de aula, relatórios de frequência e conformidade jurídica. 4) Atendimento Nacional com Atenção Local: gerentes locais que prestam suporte onsite, realizam auditorias periódicas e apoiam o alinhamento com sindicatos regionais.",
        "CHK-003": "Como política rígida de guardrails corporativos, o time de Customer Success está expressamente proibido de: 1) Prometer contratação ou preenchimento de cotas imediato em qualquer localidade (já que depende da disponibilidade de candidatos regionais homologados). 2) Garantir efetivação futura de aprendizes (visto que a decisão é exclusiva do RH e da política interna do cliente). 3) Validação jurídica própria do cliente, cabendo ao DP dele auditar as CBOs locais. 4) Disponibilidade irrestrita de turmas de formação da Leapy GO fora dos eixos de atendimento homologados.",
        "CHK-004": "Segundo o Artigo 429 da CLT e a Lei da Aprendizagem nº 10.097/2000, a contratação de jovens aprendizes é obrigatória para estabelecimentos de qualquer natureza que possuam pelo menos 7 empregados contratados in funções que exijam formação profissional. Funções administrativas, operacionais e técnicas de nível médio entram na base de cálculo. O não cumprimento da cota sujeita a empresa a multas pesadas aplicadas pela fiscalização do trabalho, além de potenciais ações civis públicas do Ministério Público do Trabalho (MPT).",
        "CHK-005": "O cálculo da cota obrigatória baseia-se na aplicação do percentual mínimo de 5% e máximo de 15% sobre o total de empregados de cada estabelecimento cujas funções demandem formação profissional, conforme a Classificação Brasileira de Ocupações (CBO). Devem ser excluídos da base de cálculo apenas os cargos de gerência, confiança, colaboradores temporários e funções que exijam nível técnico ou superior completo. A calculadora disponível na plataforma Leapy serve unicamente para simulação operacional baseada nos dados enviados pelo próprio cliente, não constituindo parecer legal.",
        "CHK-006": "A Leapy fornece orientações gerais e simulações matemáticas sobre a composição de cotas no painel do cliente. Contudo, a validação final da situação legal e da interpretação aplicável das regras oficiais de cota é de inteira responsabilidade do DP e do time jurídico do cliente. Os analistas de CS da Leapy devem recomendar que qualquer alteração de CBO ou base de cálculo simulada no painel seja previamente validada pelas assessorias legais internas do cliente antes de ser enviada ao e-Social ou reportada ao MTE.",
        "CHK-007": "Nas diretrizes oficiais de elegibilidade, os jovens aprendizes contratados sob o amparo da Lei 10.097/2000 devem ter entre 14 e 24 anos incompletos no momento da contratação. A idade máxima de 24 anos não se aplica a candidatos com deficiência (PCD), que podem ser contratados em qualquer faixa etária. O contrato de aprendizagem possui duração máxima de 24 meses, devendo ser rescindido no término do prazo ou quando o jovem completar 24 anos (exceto para PCD).",
        "CHK-008": "Para ingressar no programa de jovem aprendiz da Leapy, o candidato deve comprovar que está matriculado e frequentando regularmente o Ensino Fundamental ou Ensino Médio, ou que já concluiu a educação básica (Ensino Médio completo). A empresa contratante tem o dever de acompanhar a frequência escolar do jovem e exigir os boletins periódicos, visto que o abandono escolar ou faltas reiteradas sem justificativa na escola regular constituem motivo legal para rescisão do contrato de aprendizagem.",
        "CHK-009": "A jornada de trabalho do jovem aprendiz é rigidamente controlada por lei. Ela é limitada a no máximo 6 horas diárias para aqueles que ainda não concluíram o Ensino Fundamental (neste limite computando-se tanto as horas de atividades práticas na empresa quanto as aulas teóricas da entidade formadora). Para os jovens que já concluíram o Ensino Médio, a jornada diária pode ser de até 8 horas, contanto que haja atividades teóricas no programa de formação correspondente. São proibidas a realização de horas extras, compensações de jornada ou trabalho noturno (22h às 5h).",
        "CHK-010": "A estrutura de atendimento e operação da Leapy baseia-se em um modelo híbrido. Mantemos uma central de coordenação em São Paulo (responsável pelas diretrizes pedagógicas da Leapy GO, suporte técnico da plataforma SaaS e parametrizações de negócio) e gerentes regionais distribuídos em estados estratégicos. Isso nos permite gerenciar filiais de clientes em todo o território nacional, oferecendo proximidade física para auditorias locais, acompanhamento presencial de aprendizes e interlocução com sindicatos regionais.",
        "CHK-011": "A formação teórica do programa de jovens aprendizes deve ser realizada obrigatoriamente por entidades formadoras homologadas no CNAP (Cadastro Nacional de Aprendizagem Profissional). A Leapy utiliza a escola oficial Leapy GO para ministrar os cursos teóricos nas regiões em que possui homologação e turmas ativas. Nas localidades onde não há infraestrutura pedagógica própria da Leapy GO, realizamos parcerias com o Sistema S (Senai, Senac) ou ONGs locais parceiras.",
        "CHK-012": "Embora a plataforma SaaS da Leapy possua alcance nacional, a oferta de cursos teóricos da Leapy GO e a disponibilidade de entidades parceiras locais dependem estritamente da demanda e do cadastro municipal em cada localidade. O time de Customer Success não deve garantir a abertura de turmas presenciais imediatas para estabelecimentos do cliente situados fora dos eixos de atendimento homologados antes de confirmar a disponibilidade de vagas e homologações na região correspondente.",
        "CHK-013": "A plataforma SaaS da Leapy centraliza todas as informações do programa e indicadores de desenvolvimento dos jovens para apoiar o RH na tomada de decisão. O painel inclui controle de ponto eletrônico (frequência prática), registro de notas escolares e frequência teórica, alertas de risco de desligamento, assinaturas de contratos digitais e exportadores para o e-Social. O gestor tem visibilidade total do andamento de cada aprendiz em tempo real, eliminando o uso de planilhas paralelas.",
        "CHK-014": "A plataforma da Leapy apoia a gestão e traz visibilidade de indicadores, mas ela não substitui a atuação do departamento de Recursos Humanos do cliente. A tomada de decisões operacionais (como rescisões contratuais por justa causa, promoções de efetivação, ajustes locais de benefícios e sanções disciplinares) continua exigindo a atuação conjunta do RH da empresa e a coordenação pedagógica da Leapy.",
        "CHK-015": "A comunicação institucional pública da Leapy informa que historicamente 48% dos jovens formados pelo programa são efetivados ao fim do contrato pelas empresas contratantes. Esse índice é apresentado como três vezes maior do que a média nacional de efetivação de aprendizes, que gira em torno de 15%. Esse excelente indicador é resultado direto do nosso modelo educacional focado em capacitação técnica voltada para a economia digital (programação, dados e soft skills).",
        "CHK-016": "O time de Customer Success pode utilizar a taxa de efetivação de 48% como argumento de impacto do programa, mas deve expressamente evitar tratá-la como uma garantia contratual de resultado futuro para novas turmas de um cliente específico. A efetivação real depende da disponibilidade de vagas CLT no cliente, do orçamento anual aprovado por eles e do desempenho individual do jovem durante o contrato.",
        "CHK-017": "O assistente de inteligência operacional Leapy CSbot opera sob regras estritas de conduta e explicabilidade corporativa. Suas respostas devem ser formuladas exclusivamente com base nos fragmentos de documentos ativos no grafo de conceitos da Leapy. É terminantemente proibido inventar dados estatísticos, criar prazos de reembolso ou isenções regulatórias não homologadas, devendo o bot assumir a ausência de informações na base e ativar o fallback.",
        "CHK-018": "A estrutura padrão de entrega das respostas pelo copiloto Leapy CSbot deve respeitar obrigatoriamente os seguintes blocos de informação em formato JSON estruturado: 1) Resposta Objetiva (tonalidade profissional e direta). 2) Fontes Usadas (lista contendo doc_id e section_id dos fragmentos acessados). 3) Justificativa Curta (resumo de até 2 frases ligando as fontes à resposta). 4) Sinal de Confiança (Alta, Média, Baixa ou Nenhuma).",
        "CHK-019": "O copiloto Leapy CSbot deve ativar o mecanismo de Fallback Seguro sempre que a dúvida do usuário solicitar informações que não constam nos playbooks operacionais da empresa, tais como: precificação de pacotes e reajustes de planos comerciais, termos específicos de SLAs de contratos de parceria, integração técnica com sistemas legados ou ERPs de TI não documentados, e interpretações jurídicas e trabalhistas conclusivas.",
        "CHK-020": "Ao acionar o Fallback Seguro devido à falta de escopo ou confiança baixa nas informações documentais, o assistente deve responder com a seguinte mensagem padrão: 'Não encontrei base suficiente nos documentos disponíveis para responder com segurança. Posso indicar o que a base cobre e quais pontos exigem confirmação com o time responsável.' A justificativa do log deve descrever qual informação específica está em falta na base.",
        "CHK-021": "O dataset de avaliação (Evaluation Set) serve para auditar o comportamento da inteligência artificial frente a consultas comuns do time de Customer Success. Ele contém perguntas reais e fictícias sobre gestão, regras operacionais e restrições sindicais. As respostas geradas no protótipo devem ser auditadas manualmente no painel de auditoria, comparando-as com as respostas esperadas listadas na planilha de QA.",
        "CHK-022": "A estrutura lógica de indexação de cada fragmento documental (chunk) inserido na base de conhecimento da Leapy deve conter os seguintes campos de metadados obrigatórios: doc_id (ID do documento pai), section_id (número da seção ou parágrafo), title (título amigável da seção), audience (público receptor), topic (tópico de negócio), source_type (tipo de fonte) e updated_at (data de última revisão).",
        "CHK-023": "Para rebater a objeção do cliente 'Não quero mais um sistema para gerenciar', o CS deve argumentar que a plataforma Leapy não foi desenhada para adicionar burocracia, mas sim para simplificar e unificar processos. Ela substitui a troca caótica de e-mails, o uso de planilhas offline e o monitoramento manual de ponto por um único painel automatizado, economizando em média 12 horas semanais de trabalho do DP do cliente.",
        "CHK-024": "Frente à objeção 'Jovem aprendiz dá muito trabalho e não traz retorno', o analista deve explicar que a Leapy remove todo o esforço operacional de seleção e gestão de ponto, entregando um jovem treinado em habilidades altamente produtivas (dados e tecnologia). Isso transforma a obrigação da cota em um canal estratégico de captação de talentos de alta conversão, resultando em uma taxa de efetivação de 48%.",
        "CHK-025": "As categorias de intenção de negócio para classificação estruturada no Leapy CSbot são divididas em 8 classes nativas: empresa_visao_geral (sobre a Leapy e onboarding), cota_aprendizagem (cotas e regras legais de contratação), elegibilidade_jovem (idade, jornada e escolaridade do aprendiz), operacao_regional (cobertura nacional e acordos locais), plataforma_dados (uso do SaaS e prazos de férias), resultado_efetivacao (taxas de 48% e retenção), objecao_comercial (argumentação de vendas e integrações de TI) e fora_de_escopo (SLA, precificação ou dados contratuais privados).",
        "CHK-026": "O assistente Leapy CSbot deve associar e listar sempre de 1 a 3 fontes documentais para sustentar a resposta objetiva. Essas referências devem ser indicadas nos metadados e apresentadas na interface de forma estruturada e clicável, utilizando o formato oficial 'DOC-XXX §Y.Y'. Caso a consulta utilize mais trechos, devem ser selecionados os de maior relevância semântica.",
        "CHK-027": "Toda justificativa do assistente deve conter no máximo duas frases objetivas em português. Ela deve ligar de forma concisa o que as fontes citadas afirmam e por que aquela informação responde, limita ou nega a pergunta feita pelo analista de Customer Success, mantendo total clareza lógica para o auditor interno.",
        "CHK-028": "A classificação de 'Alta Confiança' deve ser atribuída pelo assistente somente quando o cenário da pergunta estiver inteiramente coberto pela base documental indexada, com fontes convergentes e sem ambiguidade operacional. Exemplos incluem perguntas diretas sobre idade do aprendiz, jornada de 6 horas ou e-mail corporativo da Leapy.",
        "CHK-029": "A classificação de 'Média Confiança' aplica-se a cenários onde a resposta exige cruzamento de um trecho principal com um complementar, ou quando há pequenas nuances operacionais a serem consideradas, como a obrigatoriedade de acompanhamento de frequência escolar ou a aplicação de dissídios em estados que exigem reajuste manual.",
        "CHK-030": "Caso haja qualquer conflito de informações nos playbooks ou as fontes recuperadas tragam escopo ambíguo que tenda a diminuir a confiança da resposta abaixo do nível médio, o copiloto Leapy CSbot deve recusar a resposta e acionar preventivamente o Fallback Seguro, instruindo o analista a registrar o gap para curadoria.",
        "CHK-031": "A rubrica de qualidade da auditoria de respostas baseia-se em cinco critérios de avaliação: 1) Correção Factual: aderência estrita aos documentos. 2) Rastreabilidade: citação correta das fontes. 3) Clareza: justificativa curta de até duas frases. 4) Adequação do Fallback: recusa correta em fora de escopo. 5) Utilidade: valor tático da recomendação.",
        "CHK-032": "Na conversação simulada de demonstração sobre a atuação da Leapy, o usuário pergunta se a empresa realiza apenas o curso teórico ou também faz a gestão. O assistente deve responder com alta confiança que a Leapy faz a gestão de ponta a ponta (recrutamento, contratos, frequência) e provê a formação teórica obrigatória via Leapy GO.",
        "CHK-033": "Na conversação de demonstração para testar o fallback, o analista pergunta qual é o prazo contratual de atendimento (SLA) para a empresa. Como essa informação é comercial e privada, o assistente deve acionar o Fallback Seguro, respondendo que não encontrou informações na base e sugerindo a abertura de ticket com Vendas."
    }

    doc_nodes = []
    for chunk in chunks:
        # Extract keywords
        text = chunk["chunk_text"].lower()
        title_words = re.findall(r'\w+', chunk["section_title"].lower())
        keywords = list(set([w for w in title_words if len(w) > 3] + [chunk["topic"].lower(), chunk["source_type"].lower()]))
        
        node_id = chunk["chunk_id"].lower().replace("-", "_")
        raw_text = ENRICHED_TEXTS.get(chunk["chunk_id"], chunk["chunk_text"])
        
        doc_node = {
            "id": node_id,
            "title": f"{chunk['doc_id']} §{chunk['section_id']} — {chunk['section_title']}",
            "type": "document",
            "filename": f"{chunk['doc_id'].lower()}_{chunk['section_id'].replace('.', '_')}.md",
            "topic": chunk["topic"],
            "content": raw_text.replace("'", "\\'").replace("\n", " "),
            "keywords": keywords,
            "audience": chunk["audience"],
            "source_type": chunk["source_type"],
            "updated_at": chunk["updated_at"]
        }
        doc_nodes.append(doc_node)
        
        # Determine edges to concepts based on topic/doc_id
        doc_id = chunk["doc_id"]
        # Cotas -> cota_aprendiz & cota_pcd
        if doc_id == "DOC-002":
            edges_code.append({"source": node_id, "target": "cota_aprendiz", "label": "Define"})
            edges_code.append({"source": node_id, "target": "cota_pcd", "label": "Define"})
        # Elegibilidade -> elegibilidade_clt & elegibilidade_estagio
        elif doc_id == "DOC-003":
            edges_code.append({"source": node_id, "target": "elegibilidade_estagio", "label": "Define"})
        # Operação Regional -> operacao_sudeste & operacao_nacional
        elif doc_id == "DOC-004":
            edges_code.append({"source": node_id, "target": "operacao_sudeste", "label": "Regulamenta"})
            edges_code.append({"source": node_id, "target": "operacao_nacional", "label": "Regulamenta"})
        # Plataforma -> portal_colaborador & solicitacao_ferias
        elif doc_id == "DOC-005":
            edges_code.append({"source": node_id, "target": "portal_colaborador", "label": "Explica"})
            edges_code.append({"source": node_id, "target": "solicitacao_ferias", "label": "Explica"})
        # Efetivação -> transicao_estagio
        elif doc_id == "DOC-006":
            edges_code.append({"source": node_id, "target": "transicao_estagio", "label": "Explica"})
        # Objeções -> integracao_erp
        elif doc_id == "DOC-011":
            edges_code.append({"source": node_id, "target": "integracao_erp", "label": "Responde"})
            
    # 3. Add Infrastructure files
    stack_content = ""
    if os.path.exists(".stack.md"):
        with open(".stack.md", "r", encoding="utf-8") as f:
            stack_content = f.read().replace("'", "\\'").replace("\n", " ")
            
    arch_content = ""
    if os.path.exists(".architecture.md"):
        with open(".architecture.md", "r", encoding="utf-8") as f:
            arch_content = f.read().replace("'", "\\'").replace("\n", " ")

    context_content = ""
    if os.path.exists(".context.md"):
      with open(".context.md", "r", encoding="utf-8") as f:
        context_content = f.read().replace("'", "\\'").replace("\n", " ")

    cs_content = ""
    if os.path.exists("docs/CUSTOMER_SUCCESS_LEAPY.md"):
      with open("docs/CUSTOMER_SUCCESS_LEAPY.md", "r", encoding="utf-8") as f:
        cs_content = f.read().replace("'", "\\'").replace("\n", " ")

    infra_nodes = [
        {
            "id": "doc_stack",
            "title": "DOC-017 §1.0 — Tecnologia e Stack do Leapy CSbot",
            "type": "document",
            "filename": "dotstack_recommendations.md",
            "topic": "Infraestrutura e Stack",
            "content": stack_content,
            "keywords": ["stack", "tecnologia", "react", "express", "gemini", "typescript", "vite"],
            "audience": "cs-interno",
            "source_type": "tecnico",
            "updated_at": "2026-07-19"
        },
        {
            "id": "doc_architecture",
            "title": "DOC-018 §1.0 — Arquitetura e Componentes do Leapy CSbot",
            "type": "document",
            "filename": "dotarchitecture_decisions.md",
            "topic": "Infraestrutura e Stack",
            "content": arch_content,
            "keywords": ["arquitetura", "design", "monolith", "layered", "grafo", "rag", "api"],
            "audience": "cs-interno",
            "source_type": "tecnico",
            "updated_at": "2026-07-19"
        },
        {
            "id": "doc_context",
            "title": "DOC-019 §1.0 — Contexto e Negócio do Leapy CSbot",
            "type": "document",
            "filename": "dotcontext_goals.md",
            "topic": "Infraestrutura e Stack",
            "content": context_content,
            "keywords": ["contexto", "produto", "persona", "regra", "comportamento", "fallback"],
            "audience": "cs-interno",
            "source_type": "tecnico",
            "updated_at": "2026-07-19"
        },
        {
            "id": "doc_customer_success",
            "title": "DOC-020 §1.0 — Guia de Customer Success da Leapy",
            "type": "document",
            "filename": "CUSTOMER_SUCCESS_LEAPY.md",
            "topic": "Processo e Negócio",
            "content": cs_content,
            "keywords": ["customer success", "sucesso do cliente", "onboarding", "retencao", "clientes", "publico", "leapy go"],
            "audience": "cs-interno",
            "source_type": "negocio",
            "updated_at": "2026-07-19"
        }
    ]

    # 4. Add clearly-labelled synthetic CS playbooks used by the prototype.
    # These files intentionally separate public facts from fictional scenarios.
    synthetic_specs = [
        ("01-visao-produto-e-outcomes.md", "Processo e Negócio", ["visao", "produto", "outcomes", "valor", "discovery"], ["operacao_nacional", "portal_colaborador"]),
        ("02-onboarding-implantacao.md", "Onboarding e Implantação", ["onboarding", "implantacao", "marcos", "aceite", "kickoff"], ["portal_colaborador", "operacao_nacional"]),
        ("03-integracoes-dados-seguranca.md", "Integrações e Segurança", ["integracao", "dados", "seguranca", "lgpd", "discovery"], ["integracao_erp", "seguranca_lgpd"]),
        ("04-expansao-e-objecoes.md", "Expansão e Objeções", ["expansao", "objecoes", "valor", "adocao", "oportunidade"], ["operacao_nacional", "integracao_erp"]),
        ("05-matriz-escalonamento.md", "Suporte e Escalonamento", ["escalonamento", "severidade", "triagem", "ownership", "incidente"], ["portal_colaborador", "integracao_erp"]),
        ("06-health-score-e-revisao-executiva.md", "Customer Success", ["health score", "qbr", "revisao executiva", "risco", "carteira"], ["portal_colaborador", "transicao_estagio"]),
    ]

    synthetic_nodes = []
    synthetic_dir = os.path.join("docs", "synthetic")
    for index, (filename, topic, keywords, concept_targets) in enumerate(synthetic_specs, start=1):
        path = os.path.join(synthetic_dir, filename)
        if not os.path.exists(path):
            print(f"Warning: synthetic document not found: {path}")
            continue

        with open(path, "r", encoding="utf-8") as f:
            content = f.read().strip()

        first_heading = next((line[2:].strip() for line in content.splitlines() if line.startswith("# ")), filename)
        node_id = f"doc_syn_{index:03d}"
        synthetic_nodes.append({
            "id": node_id,
            "title": first_heading,
            "type": "document",
            "filename": f"synthetic/{filename}",
            "topic": topic,
            "content": content,
            "keywords": keywords + ["demonstracao", "ficticio", "customer success"],
            "audience": "cs-interno",
            "source_type": "demonstracao-ficticia",
            "updated_at": "2026-07-19"
        })

        for concept_target in concept_targets:
            edges_code.append({"source": node_id, "target": concept_target, "label": "Demonstra"})

    # Link infra nodes to integration_erp and others
    edges_code.append({"source": "doc_stack", "target": "integracao_erp", "label": "Usa"})
    edges_code.append({"source": "doc_architecture", "target": "integracao_erp", "label": "Define"})
    edges_code.append({"source": "doc_context", "target": "portal_colaborador", "label": "Explica"})
    edges_code.append({"source": "doc_customer_success", "target": "operacao_nacional", "label": "Detona"})
    edges_code.append({"source": "doc_customer_success", "target": "portal_colaborador", "label": "Guia"})
    edges_code.append({"source": "doc_customer_success", "target": "transicao_estagio", "label": "Explica"})

    # Combine all nodes
    all_nodes = doc_nodes + infra_nodes + synthetic_nodes + concepts
    
    # Generate TypeScript code
    ts_code = """import { GraphNode, GraphEdge, DocumentNode, ConceptNode } from '../src/types.js';

// Enriched database of documents and concepts (compiled from excel and markdown docs)
export const NODES: GraphNode[] = """
    
    import json
    ts_code += json.dumps(all_nodes, indent=2, ensure_ascii=False)
    ts_code += ";\n\nexport const EDGES: GraphEdge[] = "
    ts_code += json.dumps(edges_code, indent=2, ensure_ascii=False)
    ts_code += ";\n\n"
    
    # Append retrieveWithGraph function
    with open(db_ts_path, "r", encoding="utf-8") as f:
        original_content = f.read()
        
    func_start = original_content.find("export function retrieveWithGraph")
    if func_start != -1:
        ts_code += original_content[func_start:]
    else:
        print("Error: retrieveWithGraph function not found in db.ts")
        return
        
    with open(db_ts_path, "w", encoding="utf-8") as f:
        f.write(ts_code)
        
    print("Database server/db.ts compiled successfully!")

if __name__ == "__main__":
    compile_database()
